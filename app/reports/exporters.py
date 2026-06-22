"""Reusable Excel (.xlsx) and PDF exporters for reports and list pages.

A caller builds a small data structure and hands it to `export_response`, which
returns a downloadable Flask Response in the requested format. This keeps every
export route consistent and free of openpyxl/reportlab boilerplate.

Data model — a document is a list of *blocks*. Each block is a dict::

    {
        "title": "Petrol Pump Retail",   # optional section heading
        "headers": ["Item", "Amount"],   # optional column headers
        "rows": [["Fuel sale", "123.00"], ...],  # list of list of strings
    }

Cells are expected to already be strings/numbers formatted by the caller, so the
exporter stays dumb and predictable. Use `fmt(value)` for money-style 2dp text.
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from flask import Response

VALID_FORMATS = ("xlsx", "pdf")


def fmt(value):
    """Format a number as a 2-decimal string (blank stays blank)."""
    if value is None or value == "":
        return ""
    return f"{Decimal(str(value)):.2f}"


def _stamped_name(base, ext):
    return f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def _xlsx_bytes(document_title, blocks):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(bold=True, size=14)

    row_idx = 1
    max_cols = 1

    ws.cell(row=row_idx, column=1, value=document_title).font = title_font
    row_idx += 2

    for block in blocks:
        if block.get("title"):
            ws.cell(row=row_idx, column=1, value=block["title"]).font = bold
            row_idx += 1

        headers = block.get("headers")
        if headers:
            max_cols = max(max_cols, len(headers))
            for c, head in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=c, value=head)
                cell.font = bold
                cell.fill = header_fill
            row_idx += 1

        for row in block.get("rows", []):
            max_cols = max(max_cols, len(row))
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=c, value=val)
                if c > 1:  # right-align value/amount columns
                    cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        row_idx += 1  # blank spacer between blocks

    # Reasonable column widths.
    ws.column_dimensions[get_column_letter(1)].width = 32
    for c in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #
def _pdf_bytes(document_title, blocks):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # Use landscape when any block is wide (helps list-page exports).
    wide = any(len(b.get("headers") or []) > 4 for b in blocks)
    pagesize = landscape(A4) if wide else A4

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(document_title, styles["Title"]), Spacer(1, 6 * mm)]

    for block in blocks:
        if block.get("title"):
            story.append(Paragraph(block["title"], styles["Heading3"]))

        headers = block.get("headers")
        data = ([headers] if headers else []) + [
            [str(v) for v in row] for row in block.get("rows", [])
        ]
        if not data:
            continue

        table = Table(data, repeatRows=1 if headers else 0, hAlign="LEFT")
        style = [
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Right-align everything except the first column.
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ]
        if headers:
            style += [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        table.setStyle(TableStyle(style))
        story.append(table)
        story.append(Spacer(1, 6 * mm))

    doc.build(story)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Dispatcher
# --------------------------------------------------------------------------- #
def export_response(fmt_name, base_filename, document_title, blocks):
    """Return a Flask file Response for `fmt_name` ('xlsx' or 'pdf')."""
    if fmt_name == "pdf":
        content = _pdf_bytes(document_title, blocks)
        mimetype = "application/pdf"
        filename = _stamped_name(base_filename, "pdf")
    else:  # default to xlsx
        content = _xlsx_bytes(document_title, blocks)
        mimetype = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
        filename = _stamped_name(base_filename, "xlsx")

    return Response(
        content,
        mimetype=mimetype,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
